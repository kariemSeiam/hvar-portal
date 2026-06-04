# app/api/service_api.py
"""Service API endpoints."""
from flask import Blueprint, request, jsonify, send_file, g
from app.utils.auth import require_auth
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
import logging
from app.services import service_manager, stock_manager
from app.models import service_ticket as ticket_model
from app.utils.pagination import parse_pagination_params
from app.utils.messages import get_message, get_error_message
from app.utils.phone_normalizer import normalize_ticket_phone

logger = logging.getLogger(__name__)

service_api_blueprint = Blueprint('service_api', __name__, url_prefix='/api/tickets')

# Simple in-memory cache for counts endpoint
_counts_cache = {
    'data': None,
    'timestamp': 0,
    'ttl': 8  # 8 seconds TTL
}

def invalidate_counts_cache():
    """Invalidate the counts cache. Call this when ticket status changes."""
    _counts_cache['data'] = None
    _counts_cache['timestamp'] = 0
    logger.info("[CACHE] Counts cache invalidated")

@service_api_blueprint.route('/create', methods=['POST'])
@require_auth
def create_ticket_endpoint():
    """Create a new service ticket."""
    data = request.get_json()
    ticket_type = data.get('type')
    
    # Ticket details
    customer_id = data.get('customer_id')
    user_id = g.current_user['id']
    notes = data.get('notes')
    priority = data.get('priority', 'normal')
    items = data.get('items', [])
    original_tracking = data.get('original_tracking')  # Optional
    reason = data.get('reason')
    cost_adjustment = data.get('cost_adjustment')

    # Customer details (for creation or update)
    name = data.get('name')
    phone = data.get('phone')
    phone_secondary = data.get('phone_secondary')
    city = data.get('city')
    governorate = data.get('governorate')
    address_details = data.get('address_details')

    try:
        if not customer_id and not (name and phone):
            return jsonify({"error": get_message("name_phone_required")}), 400

        if ticket_type == 'replacement':
            if not items: return jsonify({"error": get_message("items_required")}), 400
            ticket_id = service_manager.create_replacement_ticket(
                customer_id, items, user_id, notes, priority, 
                city, governorate, address_details, original_tracking,
                reason, cost_adjustment, name, phone, phone_secondary
            )
        elif ticket_type == 'maintenance':
            ticket_id = service_manager.create_maintenance_ticket(
                customer_id, user_id, notes, priority,
                reason, cost_adjustment, original_tracking, name, phone, phone_secondary,
                city, governorate, address_details, items
            )
        elif ticket_type == 'return':
            ticket_id = service_manager.create_return_ticket(
                customer_id, user_id, notes, priority,
                reason, cost_adjustment, original_tracking, name, phone, phone_secondary,
                city, governorate, address_details, items
            )
        elif ticket_type == 'sell':
            if not items: return jsonify({"error": get_message("items_required")}), 400
            customer_type = data.get('customer_type', 'customer')
            ticket_id = service_manager.create_sell_ticket(
                customer_id, items, user_id, notes, priority,
                city, governorate, address_details, original_tracking,
                reason, cost_adjustment, name, phone, phone_secondary, customer_type
            )
        else:
            return jsonify({"error": get_error_message("invalid_ticket_type", ticket_type)}), 400
        
        # Invalidate counts cache since new ticket was created
        invalidate_counts_cache()
            
        new_ticket = ticket_model.get_ticket_with_customer(ticket_id)
        new_ticket['items'] = ticket_model.get_ticket_items(ticket_id)
        new_ticket = ticket_model.enrich_ticket_with_bosta_orders(new_ticket)
        normalize_ticket_phone(new_ticket)  # Normalize phone numbers for display
        return jsonify(new_ticket), 201
    except service_manager.ServiceManagerException as e:
        return jsonify({"error": str(e)}), 400

@service_api_blueprint.route('/<int:ticket_id>/confirm', methods=['POST'])
@require_auth
def confirm_ticket_endpoint(ticket_id):
    """Confirm a service ticket, e.g., a replacement."""
    data = request.get_json()
    user_id = g.current_user['id']
    city = data.get('city')  # Optional city override
    governorate = data.get('governorate')  # Optional governorate override
    address_details = data.get('address_details')  # Optional address details override
    original_tracking = data.get('original_tracking')  # Optional original tracking (cannot be updated during confirmation)
    new_tracking_send = data.get('new_tracking_send')  # Required for replacement
    cost_adjustment = data.get('cost_adjustment', 0)  # Optional cost adjustment
    notes = data.get('notes', '')  # Optional confirmation notes
    items = data.get('items')  # Optional items to update
    phone = data.get('phone')  # Optional phone update
    phone_secondary = data.get('phone_secondary')  # Optional secondary phone update
    name = data.get('name')  # Optional customer name update
    customer_id = data.get('customer_id')  # Optional customer switching
    priority = data.get('priority')  # Optional priority update
    reason = data.get('reason')  # Optional reason update

    try:
        success = service_manager.confirm_replacement(
            ticket_id, user_id, city, governorate, address_details, original_tracking, 
            new_tracking_send, cost_adjustment, notes, items, phone, phone_secondary,
            name, customer_id, priority, reason
        )
        if success:
            # Invalidate counts cache since ticket status changed
            invalidate_counts_cache()
            
            updated_ticket = ticket_model.get_ticket_with_customer(ticket_id)
            updated_ticket['items'] = ticket_model.get_ticket_items(ticket_id)
            updated_ticket = ticket_model.enrich_ticket_with_bosta_orders(updated_ticket)
            normalize_ticket_phone(updated_ticket)  # Normalize phone numbers for display
            return jsonify(updated_ticket)
    except service_manager.ServiceManagerException as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"error": get_message("failed_execute")}), 500

@service_api_blueprint.route('/<int:ticket_id>/cancel', methods=['POST'])
@require_auth
def cancel_ticket_endpoint(ticket_id):
    """Cancel a service ticket."""
    data = request.get_json()
    user_id = g.current_user['id']
    reason = data.get('reason', get_message("ticket_cancelled"))

    try:
        success = service_manager.cancel_ticket(ticket_id, reason, user_id)
        if success:
            # Invalidate counts cache since ticket status changed
            invalidate_counts_cache()
            
            updated_ticket = ticket_model.get_ticket_with_customer(ticket_id)
            updated_ticket['items'] = ticket_model.get_ticket_items(ticket_id)
            updated_ticket = ticket_model.enrich_ticket_with_bosta_orders(updated_ticket)
            normalize_ticket_phone(updated_ticket)  # Normalize phone numbers for display
            return jsonify(updated_ticket)
    except service_manager.ServiceManagerException as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"error": get_message("err_failed_cancel")}), 500

@service_api_blueprint.route('/<int:ticket_id>', methods=['DELETE'])
@require_auth
def delete_ticket_endpoint(ticket_id):
    """Delete a cancelled service ticket permanently.

    Only tickets with CANCELLED status can be deleted.
    This operation is irreversible and removes all related data.
    """
    data = request.get_json() or {}
    user_id = g.current_user['id']

    try:
        success = service_manager.delete_ticket(ticket_id, user_id)
        if success:
            return jsonify({
                "message": get_message("deleted"),
                "ticket_id": ticket_id
            }), 200
    except service_manager.ServiceManagerException as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Unexpected error deleting ticket {ticket_id}: {str(e)}")
        return jsonify({"error": get_message("err_failed_delete_ticket")}), 500

    return jsonify({"error": get_message("err_failed_delete_ticket")}), 500


@service_api_blueprint.route('/<int:ticket_id>', methods=['GET'])
@require_auth
def get_ticket_endpoint(ticket_id):
    """Get complete ticket details with comprehensive context."""
    from app.services import tracking_manager, service_manager
    from flask import request

    ticket = ticket_model.get_ticket_with_customer(ticket_id)
    if not ticket:
        return jsonify({"error": get_message("not_found_ticket")}), 404

    # Normalize phone numbers for display
    normalize_ticket_phone(ticket)

    # Add items (fast JOIN query)
    ticket['items'] = ticket_model.get_ticket_items(ticket_id)

    # Add comprehensive scan context
    scan_history = tracking_manager.get_tracking_history_for_ticket(ticket_id)

    # Get available actions for this ticket
    available_actions = service_manager.get_available_actions(ticket_id)

    # Build warnings
    warnings = []
    if not scan_history:
        warnings.append(get_message("warning_no_history"))

    if ticket['status'] == 'CANCELLED':
        warnings.append(get_message("warning_cancelled"))

    # Add current location from last scan
    current_location = None
    if scan_history:
        last_scan = scan_history[-1]
        current_location = last_scan['scan_location']

    # Check for optional Bosta data
    include_bosta = request.args.get('include_bosta', 'auto').lower()
    if include_bosta == 'true' or (include_bosta == 'auto' and ticket.get('original_tracking')):
        ticket = ticket_model.enrich_ticket_with_bosta_orders(ticket)
    else:
        ticket['bosta_orders'] = []

    # Add scan tracking indicators if this ticket has tracking numbers
    scanned_tracking_field = None
    scanned_tracking_number = None

    # Check all tracking fields to see which one matches this ticket's data
    for field in ['original_tracking', 'new_tracking_send', 'new_tracking_receive']:
        tracking_value = ticket.get(field)
        if tracking_value:
            # This would need to be determined by how the ticket was accessed
            # For now, we'll leave it as None since we're accessing by ticket_id
            pass

    ticket['scanned_tracking_field'] = scanned_tracking_field
    ticket['scanned_tracking_number'] = scanned_tracking_number

    # Normalize phone numbers in ticket data (before building context)
    normalize_ticket_phone(ticket)

    # Build comprehensive context (same structure as scan endpoint)
    context = {
        "scan_history": scan_history,
        "ticket": ticket,
        "available_actions": available_actions,
        "warnings": warnings,
        "current_location": current_location,
        "search_summary": {
            "ticket_id": ticket['id'],
            "ticket_number": ticket['ticket_number'],
            "service_type": ticket['service_type'],
            "status": ticket['status'],
            "customer_name": ticket['customer_name'],
            "scans_count": len(scan_history),
            "has_warnings": len(warnings) > 0
        }
    }

    return jsonify({
        "ticket_id": ticket_id,
        "found": True,
        "context": context
    })

@service_api_blueprint.route('/<int:ticket_id>/history', methods=['GET'])
@require_auth
def get_ticket_history_endpoint(ticket_id):
    """Get the complete status change history for a ticket."""
    ticket = ticket_model.get_ticket_by_id(ticket_id)
    if not ticket:
        return jsonify({"error": get_message("not_found_ticket")}), 404
    
    history = ticket_model.get_ticket_history(ticket_id)
    return jsonify({
        "ticket_id": ticket_id,
        "ticket_number": ticket['ticket_number'],
        "current_status": ticket['status'],
        "history": history
    })

@service_api_blueprint.route('/', methods=['GET'])
@require_auth
def list_tickets_endpoint():
    """List all service tickets with pagination and optional filtering.

    Always returns consistent JSON format with pagination metadata:
    {
        "data": [...],
        "pagination": {
            "total": 150,
            "limit": 20,
            "offset": 0,
            "has_more": true
        }
    }

    Query Parameters:
    - service_type: replacement, maintenance, return, or sell; comma-separated for multiple
    - status: Filter by ticket status. Supports single status or comma-separated multiple statuses
              (e.g., 'CONFIRMED' or 'CONFIRMED,PENDING')
    - customer_id: Filter by specific customer
    - start_date: Filter tickets from this date (YYYY-MM-DD format)
    - end_date: Filter tickets until this date (YYYY-MM-DD format)
    - search: Search term to filter by ticket_number, customer name, phone, or tracking numbers
    - available_actions: Filter by available actions (comma-separated, e.g., 'start_preparation,ready_for_dispatch')
    - limit: Number of tickets per page (default: 20)
    - offset: Pagination offset (default: 0)
    - include_bosta: Include Bosta order data (default: true)
    - force_sync: Force fresh Bosta API sync (default: false)
    """
    limit, offset = parse_pagination_params(request)
    status = request.args.get('status')
    customer_id = request.args.get('customer_id', type=int)
    service_type = request.args.get('service_type')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    available_actions = request.args.get('available_actions')  # Comma-separated list
    force_sync = request.args.get('force_sync', 'false').lower() == 'true'
    include_bosta = request.args.get('include_bosta', 'false').lower() == 'true'

    # Validate service_type if provided (single or comma-separated, same as /filter)
    valid_service_types = ['replacement', 'maintenance', 'return', 'sell']
    if service_type:
        stypes = [s.strip() for s in service_type.split(',') if s.strip()]
        for st in stypes:
            if st not in valid_service_types:
                return jsonify({"error": get_error_message("invalid_type", st)}), 400
    
    # Validate date formats (YYYY-MM-DD format)
    if start_date:
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({"error": get_message("err_invalid_date_format")}), 400
    
    if end_date:
        try:
            datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({"error": get_message("err_invalid_date_format")}), 400

    # Determine if we have any filters
    has_filters = bool(status or customer_id or service_type or start_date or end_date or search)
    
    try:
        start_time = time.time()

        # Always use the optimized function for consistency - it handles all cases
        query_start = time.time()
        tickets = ticket_model.list_tickets_with_items_and_bosta(
            limit=limit,
            offset=offset,
            status=status,
            customer_id=customer_id,
            service_type=service_type,
            start_date=start_date,
            end_date=end_date,
            search=search,
            force_sync=force_sync,
            include_bosta=include_bosta
        )
        logger.info(f"[PERF] list_tickets_with_items_and_bosta: {(time.time() - query_start) * 1000:.2f}ms ({len(tickets)} tickets, include_bosta={include_bosta})")

        # Batch calculate available actions for all tickets (optimized - no N+1 queries)
        # Use history data already fetched in list_tickets_with_items_and_bosta
        action_start = time.time()
        history_by_ticket = {}
        for ticket in tickets:
            ticket_id = ticket['id']
            # History is already included in tickets from list_tickets_with_items_and_bosta
            history_by_ticket[ticket_id] = ticket.get('history', [])
        
        # Batch calculate all available actions at once
        actions_by_ticket = service_manager.get_available_actions_batch(tickets, history_by_ticket)
        logger.info(f"[PERF] get_available_actions_batch: {(time.time() - action_start) * 1000:.2f}ms")
        
        # Add available actions and normalize phone numbers
        normalize_start = time.time()
        for ticket in tickets:
            ticket_id = ticket['id']
            ticket['available_actions'] = actions_by_ticket.get(ticket_id, [])
            normalize_ticket_phone(ticket)  # Normalize phone numbers for display
            
            # Ensure bosta_orders exists even if empty (for consistency)
            if 'bosta_orders' not in ticket:
                ticket['bosta_orders'] = []
        logger.info(f"[PERF] Normalize phones and add actions: {(time.time() - normalize_start) * 1000:.2f}ms")

        # Filter by available_actions if specified
        if available_actions:
            filter_start = time.time()
            required_actions = [action.strip().lower() for action in available_actions.split(',')]
            filtered_tickets = []
            for ticket in tickets:
                ticket_actions = [action.lower() for action in (ticket.get('available_actions', []) or [])]
                # Ticket must have ALL specified actions
                if all(action in ticket_actions for action in required_actions):
                    filtered_tickets.append(ticket)
            tickets = filtered_tickets
            logger.info(f"[PERF] Filter by available_actions: {(time.time() - filter_start) * 1000:.2f}ms ({len(tickets)} tickets after filtering)")

        # Calculate total count based on applied filters
        count_start = time.time()
        if has_filters:
            total = ticket_model.get_tickets_count_filtered(
                status=status,
                customer_id=customer_id,
                service_type=service_type,
                start_date=start_date,
                end_date=end_date,
                search=search
            )
        else:
            total = ticket_model.get_tickets_count()
        logger.info(f"[PERF] Count query: {(time.time() - count_start) * 1000:.2f}ms (total={total})")

        # Always return consistent format with pagination metadata
        total_time = (time.time() - start_time) * 1000
        logger.info(f"[PERF] list_tickets_endpoint total time: {total_time:.2f}ms")
        return jsonify({
            "data": tickets,
            "pagination": {
                "total": total,
                "limit": limit,
                "offset": offset,
                "has_more": offset + limit < total
            }
        })
    
    except ConnectionError as e:
        logger.error(f"Database connection error in list_tickets_endpoint: {str(e)}")
        return jsonify({
            "error": "Database connection failed",
            "message": "Unable to connect to the database. Please ensure MySQL is running.",
            "details": str(e),
            "data": [],
            "pagination": {
                "total": 0,
                "limit": limit,
                "offset": offset,
                "has_more": False
            }
        }), 503  # Service Unavailable
    
    except Exception as e:
        logger.error(f"Unexpected error in list_tickets_endpoint: {str(e)}", exc_info=True)
        return jsonify({
            "error": "Internal server error",
            "message": "An unexpected error occurred while fetching tickets.",
            "details": str(e),
            "data": [],
            "pagination": {
                "total": 0,
                "limit": limit,
                "offset": offset,
                "has_more": False
            }
        }), 500

@service_api_blueprint.route('/counts', methods=['GET'])
@require_auth
def get_ticket_counts_endpoint():
    """Get ticket counts for all tabs and sub-tabs efficiently.

    Returns counts grouped by service_type and sub-tab ID, optimized for hub page display.
    Uses efficient SQL queries to count ALL tickets (not limited to 1000).
    Uses in-memory caching with 8 second TTL to reduce database load.
    """
    from app.utils.db import execute_query
    
    try:
        # Check cache first
        current_time = time.time()
        force_refresh = request.args.get('force_refresh', 'false').lower() == 'true'
        
        if not force_refresh and _counts_cache['data'] and (current_time - _counts_cache['timestamp']) < _counts_cache['ttl']:
            logger.info(f"[PERF] get_ticket_counts_endpoint: Cache HIT (age: {(current_time - _counts_cache['timestamp']) * 1000:.2f}ms)")
            return jsonify(_counts_cache['data'])
        
        logger.info(f"[PERF] get_ticket_counts_endpoint: Cache MISS - fetching from database")
        
        start_time = time.time()
        counts = {
            'replacement': {},
            'maintenance': {},
            'return': {},
            'sell': {},
            'pending': 0,
        }

        # ========== REPLACEMENT COUNTS ==========
        # Simple status-based counts
        query_start = time.time()
        replacement_simple_counts = execute_query("""
            SELECT
                status,
                COUNT(*) as count
            FROM service_tickets
            WHERE service_type = 'replacement'
            GROUP BY status
        """)
        logger.info(f"[PERF] Replacement simple counts query: {(time.time() - query_start) * 1000:.2f}ms")

        for row in replacement_simple_counts:
            status = row['status']
            count = row['count']

            # Map statuses to sub-tab IDs
            if status == 'READY_FOR_DISPATCH':
                counts['replacement']['ready-to-ship'] = count
            elif status == 'SENT':
                counts['replacement']['sent'] = count
            elif status == 'RETURNED':
                counts['replacement']['validate-returns'] = count
            elif status == 'COMPLETED':
                counts['replacement']['completed'] = count
            elif status == 'CANCELLED':
                counts['replacement']['cancelled'] = count

        # in-preparation: CONFIRMED replacement tickets (all have 'start_preparation' action)
        # preparing: IN_PROCESS replacement tickets (all have 'ready_for_dispatch' action)
        # Optimized: Single SQL query instead of N+1 queries
        query_start = time.time()
        replacement_action_counts = execute_query("""
            SELECT 
                COUNT(CASE WHEN status = 'CONFIRMED' THEN 1 END) as in_preparation,
                COUNT(CASE WHEN status = 'IN_PROCESS' THEN 1 END) as preparing
            FROM service_tickets
            WHERE service_type = 'replacement' 
              AND status IN ('CONFIRMED', 'IN_PROCESS')
        """)
        logger.info(f"[PERF] Replacement action counts query: {(time.time() - query_start) * 1000:.2f}ms")
        counts['replacement']['in-preparation'] = replacement_action_counts[0]['in_preparation'] if replacement_action_counts else 0
        counts['replacement']['preparing'] = replacement_action_counts[0]['preparing'] if replacement_action_counts else 0

        # ========== MAINTENANCE COUNTS ==========
        # Simple status-based counts
        query_start = time.time()
        maintenance_simple_counts = execute_query("""
            SELECT
                status,
                COUNT(*) as count
            FROM service_tickets
            WHERE service_type = 'maintenance'
            GROUP BY status
        """)
        logger.info(f"[PERF] Maintenance simple counts query: {(time.time() - query_start) * 1000:.2f}ms")

        for row in maintenance_simple_counts:
            status = row['status']
            count = row['count']

            # Map statuses to sub-tab IDs
            if status == 'READY_FOR_DISPATCH':
                counts['maintenance']['ready-to-ship'] = count
            elif status == 'SENT':
                counts['maintenance']['sent'] = count
            elif status == 'COMPLETED':
                counts['maintenance']['completed'] = count
            elif status == 'CANCELLED':
                counts['maintenance']['cancelled'] = count

        # Maintenance: split CONFIRMED vs PENDING so frontend can:
        # - Main "مؤكد" status tab badge = CONFIRMED-only (matches listTickets status=CONFIRMED)
        # - Sub-tab "مؤكد" under الصيانة = confirmed + pending (matches list CONFIRMED,PENDING)
        query_start = time.time()
        maint_conf = execute_query("""
            SELECT COUNT(*) as count
            FROM service_tickets
            WHERE service_type = 'maintenance' AND status = 'CONFIRMED'
        """)
        maint_pend = execute_query("""
            SELECT COUNT(*) as count
            FROM service_tickets
            WHERE service_type = 'maintenance' AND status = 'PENDING'
        """)
        counts['maintenance']['confirmed'] = maint_conf[0]['count'] if maint_conf else 0
        counts['maintenance']['pending'] = maint_pend[0]['count'] if maint_pend else 0
        logger.info(f"[PERF] Maintenance confirmed/pending split query: {(time.time() - query_start) * 1000:.2f}ms")

        # For IN_PROCESS maintenance tickets, we need to count history entries to determine the state
        # received: IN_PROCESS + 0 internal actions (start_maintenance action available)
        # under-maintenance: IN_PROCESS + 1 internal action (complete_maintenance action available)
        # completion-ready: IN_PROCESS + 2+ internal actions (mark_ready action available)
        # Optimized: Single SQL query with conditional aggregation instead of Python loop
        query_start = time.time()
        maintenance_in_process_counts = execute_query("""
            SELECT
                COUNT(CASE
                    WHEN internal_action_count = 0 THEN 1
                END) as received,
                COUNT(CASE
                    WHEN internal_action_count = 1 THEN 1
                END) as under_maintenance,
                COUNT(CASE
                    WHEN internal_action_count >= 2 THEN 1
                END) as completion_ready
            FROM (
                SELECT
                    st.id,
                    COUNT(CASE
                        WHEN sth.old_status = 'IN_PROCESS' AND sth.new_status = 'IN_PROCESS'
                        THEN 1
                    END) as internal_action_count
                FROM service_tickets st
                LEFT JOIN service_ticket_history sth ON st.id = sth.ticket_id
                WHERE st.service_type = 'maintenance' AND st.status = 'IN_PROCESS'
                GROUP BY st.id
            ) as ticket_counts
        """)
        query_time = (time.time() - query_start) * 1000
        logger.info(f"[PERF] Maintenance IN_PROCESS counts query (optimized): {query_time:.2f}ms")
        
        if maintenance_in_process_counts and len(maintenance_in_process_counts) > 0:
            counts['maintenance']['received'] = maintenance_in_process_counts[0].get('received', 0) or 0
            counts['maintenance']['under-maintenance'] = maintenance_in_process_counts[0].get('under_maintenance', 0) or 0
            counts['maintenance']['completion-ready'] = maintenance_in_process_counts[0].get('completion_ready', 0) or 0
        else:
            counts['maintenance']['received'] = 0
            counts['maintenance']['under-maintenance'] = 0
            counts['maintenance']['completion-ready'] = 0

        # ========== RETURN COUNTS ==========
        # Simple status-based counts
        query_start = time.time()
        return_simple_counts = execute_query("""
            SELECT
                status,
                COUNT(*) as count
            FROM service_tickets
            WHERE service_type = 'return'
            GROUP BY status
        """)
        logger.info(f"[PERF] Return simple counts query: {(time.time() - query_start) * 1000:.2f}ms")

        for row in return_simple_counts:
            status = row['status']
            count = row['count']

            # Map statuses to sub-tab IDs
            if status == 'CONFIRMED':
                counts['return']['receiving'] = count
            elif status == 'IN_PROCESS':
                counts['return']['inspection'] = count
            elif status == 'COMPLETED':
                counts['return']['completed'] = count
            elif status == 'CANCELLED':
                counts['return']['cancelled'] = count

        # ========== SELL COUNTS ==========
        # Simple status-based counts
        query_start = time.time()
        sell_simple_counts = execute_query("""
            SELECT
                status,
                COUNT(*) as count
            FROM service_tickets
            WHERE service_type = 'sell'
            GROUP BY status
        """)
        logger.info(f"[PERF] Sell simple counts query: {(time.time() - query_start) * 1000:.2f}ms")

        for row in sell_simple_counts:
            status = row['status']
            count = row['count']

            # Map statuses to sub-tab IDs
            if status == 'CONFIRMED':
                counts['sell']['new'] = count  # 'new' tab shows CONFIRMED tickets
            elif status == 'READY_FOR_DISPATCH':
                counts['sell']['ready-to-ship'] = count
            elif status == 'SENT':
                counts['sell']['sent'] = count
            elif status == 'COMPLETED':
                counts['sell']['completed'] = count
            elif status == 'CANCELLED':
                counts['sell']['cancelled'] = count
            elif status == 'RETURNED':
                counts['sell']['returned'] = count

        # preparing: IN_PROCESS sell tickets (all have 'ready_for_dispatch' action)
        # Optimized: Single SQL COUNT query instead of N+1 queries
        query_start = time.time()
        sell_preparing_result = execute_query("""
            SELECT COUNT(*) as count
            FROM service_tickets
            WHERE service_type = 'sell' AND status = 'IN_PROCESS'
        """)
        logger.info(f"[PERF] Sell preparing count query: {(time.time() - query_start) * 1000:.2f}ms")
        counts['sell']['preparing'] = sell_preparing_result[0]['count'] if sell_preparing_result else 0

        # Ensure all sub-tabs have a count (default to 0 if not set)
        replacement_subtabs = ['in-preparation', 'preparing', 'ready-to-ship', 'sent', 'validate-returns', 'completed', 'cancelled']
        maintenance_subtabs = ['confirmed', 'pending', 'received', 'under-maintenance', 'completion-ready', 'ready-to-ship', 'sent', 'completed', 'cancelled']
        return_subtabs = ['receiving', 'inspection', 'completed', 'cancelled']
        sell_subtabs = ['new', 'preparing', 'ready-to-ship', 'sent', 'returned', 'completed', 'cancelled']

        for subtab in replacement_subtabs:
            if subtab not in counts['replacement']:
                counts['replacement'][subtab] = 0

        for subtab in maintenance_subtabs:
            if subtab not in counts['maintenance']:
                counts['maintenance'][subtab] = 0

        for subtab in return_subtabs:
            if subtab not in counts['return']:
                counts['return'][subtab] = 0

        for subtab in sell_subtabs:
            if subtab not in counts['sell']:
                counts['sell'][subtab] = 0

        # Global PENDING count for انتظار tab (tickets shown as "قيد الانتظار" in table)
        query_start = time.time()
        pending_result = execute_query("""
            SELECT COUNT(*) AS `count` FROM service_tickets
            WHERE status IN ('PENDING', 'pending')
        """)
        logger.info(f"[PERF] PENDING count query: {(time.time() - query_start) * 1000:.2f}ms")
        if pending_result and len(pending_result) > 0:
            row = pending_result[0]
            # DictCursor may return key as 'count' or 'COUNT(*)'; support both
            counts['pending'] = int(row.get('count', row.get('COUNT(*)', 0)) or 0)
        else:
            counts['pending'] = 0

        total_time = (time.time() - start_time) * 1000
        logger.info(f"[PERF] get_ticket_counts_endpoint total time: {total_time:.2f}ms")
        
        # Update cache
        _counts_cache['data'] = counts
        _counts_cache['timestamp'] = time.time()
        
        return jsonify(counts)
    
    except ConnectionError as e:
        logger.error(f"Database connection error in get_ticket_counts_endpoint: {str(e)}")
        return jsonify({
            "error": "Database connection failed",
            "message": "Unable to connect to the database. Please ensure MySQL is running.",
            "details": str(e),
            "replacement": {},
            "maintenance": {},
            "return": {},
            "sell": {},
            "pending": 0,
        }), 503  # Service Unavailable
    
    except Exception as e:
        logger.error(f"Unexpected error in get_ticket_counts_endpoint: {str(e)}", exc_info=True)
        return jsonify({
            "error": "Internal server error",
            "message": "An unexpected error occurred while fetching ticket counts.",
            "details": str(e),
            "replacement": {},
            "maintenance": {},
            "return": {},
            "sell": {},
            "pending": 0,
        }), 500

@service_api_blueprint.route('/<int:ticket_id>/actions', methods=['GET'])
@require_auth
def get_actions_endpoint(ticket_id):
    """Get available actions for a ticket."""
    actions = service_manager.get_available_actions(ticket_id)
    return jsonify({"available_actions": actions})

@service_api_blueprint.route('/<int:ticket_id>/action', methods=['POST'])
@require_auth
def execute_action_endpoint(ticket_id):
    """Execute a workflow action on a ticket."""
    data = request.get_json()
    action = data.get('action')
    user_id = g.current_user['id']
    notes = data.get('notes', '')  # Optional notes for each action
    cost_adjustment = data.get('cost_adjustment', 0)  # Optional cost adjustment
    items = data.get('items')  # Optional items to update (for confirm action)
    phone = data.get('phone')  # Optional phone update (for confirm action)
    phone_secondary = data.get('phone_secondary')  # Optional secondary phone update (for confirm action)
    name = data.get('name')  # Optional customer name update (for confirm action)
    customer_id = data.get('customer_id')  # Optional customer switching (for confirm action)
    priority = data.get('priority')  # Optional priority update (for confirm action)
    reason = data.get('reason')  # Optional reason update (for confirm action)
    
    if not action:
        return jsonify({"error": get_message("missing_fields")}), 400
    
    try:
        if action == 'confirm':
            # Get ticket to determine service type
            ticket = ticket_model.get_ticket_by_id(ticket_id)
            if not ticket:
                return jsonify({"error": get_message("not_found_ticket")}), 404
                
            service_type = ticket['service_type']
            
            if service_type == 'replacement':
                city = data.get('city')
                governorate = data.get('governorate')
                address_details = data.get('address_details')
                original_tracking = data.get('original_tracking')
                new_tracking_send = data.get('new_tracking_send')
                new_tracking_receive = data.get('new_tracking_receive')
                if not new_tracking_send:
                    return jsonify({"error": get_message("missing_tracking")}), 400
                success = service_manager.confirm_replacement(
                    ticket_id, user_id, city, governorate, address_details, original_tracking,
                    new_tracking_send, new_tracking_receive, cost_adjustment, notes, items, phone, phone_secondary,
                    name, customer_id, priority, reason
                )
            elif service_type == 'maintenance':
                new_tracking_send = data.get('new_tracking_send')
                new_tracking_receive = data.get('new_tracking_receive')
                success = service_manager.confirm_maintenance(
                    ticket_id, user_id, new_tracking_send, new_tracking_receive, cost_adjustment, notes, items, phone, phone_secondary,
                    name, customer_id, priority, reason
                )
            elif service_type == 'return':
                city = data.get('city')
                governorate = data.get('governorate')
                address_details = data.get('address_details')
                new_tracking_send = data.get('new_tracking_send')
                new_tracking_receive = data.get('new_tracking_receive')
                success = service_manager.confirm_return(
                    ticket_id, user_id, new_tracking_send, new_tracking_receive, cost_adjustment, notes, items, phone, phone_secondary,
                    name, customer_id, priority, reason, city, governorate, address_details
                )
            elif service_type == 'sell':
                city = data.get('city')
                governorate = data.get('governorate')
                address_details = data.get('address_details')
                original_tracking = data.get('original_tracking')
                new_tracking_send = data.get('new_tracking_send')
                new_tracking_receive = data.get('new_tracking_receive')
                success = service_manager.confirm_sell(
                    ticket_id, user_id, city, governorate, address_details, original_tracking,
                    new_tracking_send, new_tracking_receive, cost_adjustment, notes, items, phone, phone_secondary,
                    name, customer_id, priority, reason
                )
            else:
                return jsonify({"error": get_error_message("invalid_type", service_type)}), 400
        elif action == 'start_preparation':
            success = service_manager.start_preparation(ticket_id, user_id, notes, cost_adjustment)
        elif action == 'start_maintenance':
            success = service_manager.start_maintenance(ticket_id, user_id, notes)
        elif action == 'complete_maintenance':
            # Items are now optional - complete_maintenance handles None/empty items
            success = service_manager.complete_maintenance(ticket_id, user_id, notes, cost_adjustment, items)
        elif action == 'mark_ready':
            new_tracking_send = data.get('new_tracking_send')
            if not new_tracking_send:
                return jsonify({"error": get_message("missing_tracking")}), 400
            success = service_manager.mark_ready(ticket_id, new_tracking_send, user_id, notes, cost_adjustment)
        elif action == 'ready_for_dispatch':
            success = service_manager.ready_for_dispatch(ticket_id, user_id, notes, cost_adjustment)
        elif action == 'scan_outbound':
            tracking_number = data.get('tracking_number')
            if not tracking_number:
                return jsonify({"error": get_message("missing_tracking")}), 400
            success = service_manager.scan_outbound(ticket_id, tracking_number, user_id, notes, cost_adjustment)
        elif action == 'scan_inbound':
            tracking_number = data.get('tracking_number')
            if not tracking_number:
                return jsonify({"error": get_message("missing_tracking")}), 400
            success = service_manager.scan_inbound(ticket_id, tracking_number, user_id, notes, cost_adjustment)
        elif action == 'validate_items':
            item_validations = data.get('item_validations', [])
            if not item_validations:
                return jsonify({"error": get_message("missing_fields")}), 400
            success = service_manager.validate_items(ticket_id, item_validations, user_id, notes, cost_adjustment)
        elif action == 'mark_delivered':
            item_validations = data.get('item_validations')
            success = service_manager.mark_delivered(ticket_id, user_id, notes, cost_adjustment, item_validations)
        elif action == 'complete':
            success = service_manager.complete_ticket(ticket_id, user_id, notes)
        elif action == 'confirm_sent':
            # For sell tickets: confirm delivery and mark as completed (similar to mark_delivered for maintenance)
            item_validations = data.get('item_validations')
            success = service_manager.confirm_sent(ticket_id, user_id, notes, cost_adjustment, item_validations)
        else:
            return jsonify({"error": get_error_message("invalid_type", action)}), 400
            
        if success:
            # Invalidate counts cache since ticket status/actions changed
            invalidate_counts_cache()
            
            # Fetch updated ticket with all current data
            updated_ticket = ticket_model.get_ticket_with_customer(ticket_id)
            updated_ticket['items'] = ticket_model.get_ticket_items(ticket_id)
            updated_ticket = ticket_model.enrich_ticket_with_bosta_orders(updated_ticket)
            # Add available_actions to response so frontend can immediately show correct actions
            updated_ticket['available_actions'] = service_manager.get_available_actions(ticket_id)
            normalize_ticket_phone(updated_ticket)  # Normalize phone numbers for display
            return jsonify(updated_ticket)
        else:
            return jsonify({"error": get_message("failed_execute")}), 500
            
    except stock_manager.StockManagerException as e:
        return jsonify({"error": str(e)}), 400
    except service_manager.ServiceManagerException as e:
        return jsonify({"error": str(e)}), 400

@service_api_blueprint.route('/filter', methods=['GET'])
@require_auth
def filter_tickets_endpoint():
    """Get filter summary counts OR export filtered tickets to Excel.
    
    Query Parameters (same as list_tickets_endpoint):
    - service_type: Filter by ticket type (replacement, maintenance, return, sell) or comma-separated
    - status: Comma-separated statuses (e.g., 'CONFIRMED,PENDING')
    - start_date: Filter from date (YYYY-MM-DD)
    - end_date: Filter until date (YYYY-MM-DD)
    - search: Search term
    - export: If 'true', returns Excel file instead of JSON summary
    
    Returns:
    - If export=true: Excel file download
    - Otherwise: JSON summary with total, byServiceType, byStatus counts
    """
    try:
        # Get filter parameters
        status = request.args.get('status')
        service_type = request.args.get('service_type')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        search = request.args.get('search')
        export = request.args.get('export', 'false').lower() == 'true'
        
        # Validate service_type if provided
        valid_service_types = ['replacement', 'maintenance', 'return', 'sell']
        if service_type:
            service_types = [s.strip() for s in service_type.split(',')] if ',' in service_type else [service_type]
            for st in service_types:
                if st not in valid_service_types:
                    return jsonify({"error": get_error_message("invalid_type", st)}), 400
        
        # Validate date formats
        if start_date:
            try:
                datetime.strptime(start_date, '%Y-%m-%d')
            except ValueError:
                return jsonify({"error": get_message("err_invalid_date_format")}), 400
        
        if end_date:
            try:
                datetime.strptime(end_date, '%Y-%m-%d')
            except ValueError:
                return jsonify({"error": get_message("err_invalid_date_format")}), 400
        
        # Handle Excel export
        if export:
            # Fetch ALL matching tickets (no pagination limit)
            # Handle multiple service types by fetching separately and combining
            all_tickets = []
            
            if service_type and ',' in service_type:
                # Multiple service types: fetch each separately
                service_types = [s.strip() for s in service_type.split(',')]
                for st in service_types:
                    tickets = ticket_model.list_tickets_with_items_and_bosta(
                        limit=10000,  # Large limit to get all tickets
                        offset=0,
                        status=status,
                        service_type=st,
                        start_date=start_date,
                        end_date=end_date,
                        search=search,
                        force_sync=False,
                        include_bosta=False  # Don't need Bosta data for export
                    )
                    all_tickets.extend(tickets)
            else:
                # Single service type or no service type filter
                all_tickets = ticket_model.list_tickets_with_items_and_bosta(
                    limit=10000,  # Large limit to get all tickets
                    offset=0,
                    status=status,
                    service_type=service_type,
                    start_date=start_date,
                    end_date=end_date,
                    search=search,
                    force_sync=False,
                    include_bosta=False  # Don't need Bosta data for export
                )
            
            if not all_tickets:
                return jsonify({"error": "لا توجد بيانات للتصدير"}), 404
            
            tickets = all_tickets
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "التذاكر"
            
            # Define header style
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12, name='Arial')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Arabic headers
            headers = ['نوع الخدمة', 'رقم التذكرة', 'العميل', 'الهاتف', 'الحالة', 'المنتجات المرسلة', 'المنتجات المستلمة', 'أرقام التتبع', 'التاريخ', 'الوقت', 'الملاحظات']
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Service type labels in Arabic
            service_type_labels = {
                'replacement': 'استبدال',
                'maintenance': 'صيانة',
                'return': 'استرجاع',
                'sell': 'المبيعات'
            }
            
            # Status labels in Arabic
            status_labels = {
                'PENDING': 'قيد الانتظار',
                'CONFIRMED': 'مؤكد',
                'IN_PROCESS': 'قيد المعالجة',
                'READY_FOR_DISPATCH': 'جاهز للإرسال',
                'SENT': 'مرسل',
                'DELIVERED': 'تم التسليم',
                'RETURNED': 'مرتجع',
                'COMPLETED': 'مكتمل',
                'CANCELLED': 'ملغي'
            }
            
            # Helper function to format items by direction
            def format_items_by_direction(items, direction):
                """Format items for a specific direction (send/receive).
                
                Format: Item Name (Quantity) - SKU - Condition
                Each item on a new line.
                """
                if not items:
                    return ''
                
                filtered_items = [item for item in items if item.get('direction') == direction]
                if not filtered_items:
                    return ''
                
                formatted_lines = []
                for item in filtered_items:
                    name = item.get('name', item.get('item_name', ''))
                    quantity = item.get('quantity', 0)
                    sku = item.get('sku', '')
                    condition = item.get('condition', '')
                    
                    # Translate condition to Arabic
                    condition_labels = {
                        'valid': 'صالح',
                        'damaged': 'تالف'
                    }
                    condition_ar = condition_labels.get(condition, condition)
                    
                    # Format: Name (Qty) - SKU - Condition
                    line = f"{name} ({quantity}) - {sku} - {condition_ar}"
                    formatted_lines.append(line)
                
                return '\n'.join(formatted_lines)
            
            # Helper function to format tracking numbers
            def format_tracking_numbers(send_tracking, receive_tracking):
                """Format tracking numbers for send and receive.
                
                If both exist and are the same, show both labels.
                If different, show both separately.
                If one is empty, show only the available one.
                """
                send_tracking = send_tracking or ''
                receive_tracking = receive_tracking or ''
                
                if not send_tracking and not receive_tracking:
                    return ''
                
                if send_tracking and receive_tracking:
                    if send_tracking == receive_tracking:
                        # Same tracking number, show both labels
                        return f"إرسال: {send_tracking}\nاستلام: {receive_tracking}"
                    else:
                        # Different tracking numbers
                        return f"إرسال: {send_tracking}\nاستلام: {receive_tracking}"
                elif send_tracking:
                    return f"إرسال: {send_tracking}"
                elif receive_tracking:
                    return f"استلام: {receive_tracking}"
                
                return ''
            
            # Write data rows
            for row_num, ticket in enumerate(tickets, 2):
                # Format send items
                send_items_str = format_items_by_direction(ticket.get('items', []), 'send')
                
                # Format receive items
                receive_items_str = format_items_by_direction(ticket.get('items', []), 'receive')
                
                # Format tracking numbers
                tracking_str = format_tracking_numbers(
                    ticket.get('new_tracking_send'),
                    ticket.get('new_tracking_receive')
                )
                
                # Format date and time
                created_at = ticket.get('created_at')
                date_str = ''
                time_str = ''
                if created_at:
                    try:
                        dt = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S') if isinstance(created_at, str) else created_at
                        date_str = dt.strftime('%Y-%m-%d')
                        time_str = dt.strftime('%H:%M:%S')
                    except:
                        date_str = str(created_at)[:10] if created_at else ''
                        time_str = str(created_at)[11:19] if created_at and len(str(created_at)) > 10 else ''
                
                # Get customer info
                customer_name = ticket.get('customer_name', ticket.get('name', ''))
                customer_phone = ticket.get('customer_phone', ticket.get('phone', ''))
                
                # Write row data (REMOVED original_tracking, ADDED send_items, receive_items, tracking)
                row_data = [
                    service_type_labels.get(ticket.get('service_type', ''), ticket.get('service_type', '')),
                    ticket.get('ticket_number', ''),
                    customer_name,
                    customer_phone,
                    status_labels.get(ticket.get('status', ''), ticket.get('status', '')),
                    send_items_str,        # NEW: Send items column
                    receive_items_str,     # NEW: Receive items column
                    tracking_str,          # NEW: Tracking numbers column
                    date_str,
                    time_str,
                    ticket.get('notes', '')
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    # For multi-line cells (items and tracking), use left alignment and enable wrap text
                    if col_num in [6, 7, 8]:  # Send items, Receive items, Tracking columns
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15  # نوع الخدمة
            ws.column_dimensions['B'].width = 20  # رقم التذكرة
            ws.column_dimensions['C'].width = 30  # العميل
            ws.column_dimensions['D'].width = 15  # الهاتف
            ws.column_dimensions['E'].width = 18  # الحالة
            ws.column_dimensions['F'].width = 45  # المنتجات المرسلة (Send Items - wider for multi-line)
            ws.column_dimensions['G'].width = 45  # المنتجات المستلمة (Receive Items - wider for multi-line)
            ws.column_dimensions['H'].width = 25  # أرقام التتبع (Tracking Numbers)
            ws.column_dimensions['I'].width = 12  # التاريخ
            ws.column_dimensions['J'].width = 10  # الوقت
            ws.column_dimensions['K'].width = 50  # الملاحظات (Notes - no strict limit)
            
            # Set right-to-left direction for the sheet
            ws.sheet_view.rightToLeft = False
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"tickets_export_{timestamp}.xlsx"
            
            # Return file
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        
        # Handle JSON summary (default behavior)
        # Build base filters (excluding service_type for breakdown calculation)
        base_filters = {
            'status': status,
            'start_date': start_date,
            'end_date': end_date,
            'search': search
        }
        
        # Calculate total count
        if service_type:
            # If single service type, use it in count
            if ',' not in service_type:
                total = ticket_model.get_tickets_count_filtered(
                    status=status,
                    service_type=service_type,
                    start_date=start_date,
                    end_date=end_date,
                    search=search
                )
            else:
                # Multiple service types: sum counts for each
                service_types = [s.strip() for s in service_type.split(',')]
                total = sum(
                    ticket_model.get_tickets_count_filtered(
                        status=status,
                        service_type=st,
                        start_date=start_date,
                        end_date=end_date,
                        search=search
                    )
                    for st in service_types
                )
        else:
            # No service type filter: count all
            total = ticket_model.get_tickets_count_filtered(
                status=status,
                start_date=start_date,
                end_date=end_date,
                search=search
            )
        
        # Calculate byServiceType breakdown
        by_service_type = {}
        service_types_to_check = (
            [s.strip() for s in service_type.split(',')]
            if service_type
            else ['replacement', 'maintenance', 'return', 'sell']
        )
        
        for st in service_types_to_check:
            count = ticket_model.get_tickets_count_filtered(
                status=status,
                service_type=st,
                start_date=start_date,
                end_date=end_date,
                search=search
            )
            by_service_type[st] = count
        
        # Calculate byStatus breakdown (only if status filter is active)
        by_status = None
        if status:
            by_status = {}
            statuses = [s.strip() for s in status.split(',')] if ',' in status else [status]
            
            for stat in statuses:
                count = ticket_model.get_tickets_count_filtered(
                    status=stat,
                    service_type=service_type,
                    start_date=start_date,
                    end_date=end_date,
                    search=search
                )
                by_status[stat] = count
        
        # Return JSON summary
        return jsonify({
            "total": total,
            "byServiceType": by_service_type,
            "byStatus": by_status
        })
        
    except Exception as e:
        print(f"Filter endpoint error: {str(e)}")
        return jsonify({"error": "فشل في معالجة الطلب"}), 500
