# app/api/stock_api.py
"""Stock API endpoints."""
import logging
from flask import Blueprint, request, jsonify, send_file, g
from app.utils.auth import require_auth
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services import stock_manager
from app.models import stock as stock_model
from app.utils.db import execute_insert, execute_update
from app.utils.messages import get_message, get_error_message

logger = logging.getLogger(__name__)

stock_api_blueprint = Blueprint('stock_api', __name__, url_prefix='/api/stock')

def _add_component_to_product(product_id, part_id, quantity_needed):
    """
    Helper function to add a component to a product.
    Validates the part and adds it to product_components table.
    Raises ValueError if validation fails.
    """
    # Verify part exists, is active, and is a part
    part = stock_model.get_stock_item_by_id(part_id, active_only=True)
    if not part:
        raise ValueError(f"Part with ID {part_id} not found or is inactive")
    
    if part['type'] != 'part':
        raise ValueError(f"Item with ID {part_id} is not a part")
    
    # Add component
    component_sql = """
        INSERT INTO product_components (product_id, part_id, quantity_needed)
        VALUES (%s, %s, %s)
    """
    execute_insert(component_sql, (product_id, part_id, quantity_needed))

@stock_api_blueprint.route('/items', methods=['POST'])
@require_auth
def create_item():
    """Create a new stock item (part or product)."""
    data = request.get_json()
    sku = data.get('sku')
    name = data.get('name')
    item_type = data.get('type')
    quantity_on_hand = data.get('quantity_on_hand', 0)
    user_id = g.current_user['id']
    price_customer = data.get('price_customer')
    price_merchant = data.get('price_merchant')
    components = data.get('components', [])  # For products: [{"part_id": 1, "quantity_needed": 2}, ...]

    # Validate required fields
    if not all([sku, name, item_type]):
        return jsonify({"error": get_message("missing_fields")}), 400

    # Validate item type
    if item_type not in ['part', 'product']:
        return jsonify({"error": get_message("invalid_item_type")}), 400

    # Check if SKU already exists
    existing = stock_model.get_stock_item_by_sku(sku)
    if existing:
        return jsonify({"error": get_message("duplicate_sku")}), 409

    try:
        # Create the stock item
        item_data = {
            'sku': sku,
            'name': name,
            'type': item_type,
            'quantity_on_hand': quantity_on_hand,
            'created_by': user_id
        }
        if price_customer is not None:
            item_data['price_customer'] = price_customer
        if price_merchant is not None:
            item_data['price_merchant'] = price_merchant
        item_id = stock_model.create_stock_item(item_data)

        # If it's a product with components, add them
        if item_type == 'product' and components:
            for component in components:
                part_id = component.get('part_id')
                quantity_needed = component.get('quantity_needed')
                
                if not part_id or not quantity_needed:
                    return jsonify({"error": get_message("missing_fields")}), 400
                
                try:
                    _add_component_to_product(item_id, part_id, quantity_needed)
                except ValueError as e:
                    return jsonify({"error": get_message("not_found_item")}), 404

        # Get the created item with all details
        created_item = stock_model.get_stock_item_by_id(item_id)
        
        # If product, include components
        if item_type == 'product':
            created_item['components'] = stock_model.get_product_components(item_id)
        
        return jsonify(created_item), 201

    except Exception as e:
        return jsonify({"error": get_message("failed_create")}), 500

@stock_api_blueprint.route('/items/<int:item_id>', methods=['PUT'])
@require_auth
def update_item(item_id):
    """Update stock item details."""
    data = request.get_json()
    sku = data.get('sku')
    name = data.get('name')
    active = data.get('active')
    user_id = data.get('user_id')
    price_customer = data.get('price_customer')
    price_merchant = data.get('price_merchant')

    # Validate at least one field to update
    if not any([sku, name, active is not None, user_id, price_customer is not None, price_merchant is not None]):
        return jsonify({"error": get_message("no_changes")}), 400

    # Check if item exists
    item = stock_model.get_stock_item_by_id(item_id)
    if not item:
        return jsonify({"error": get_message("not_found_stock")}), 404

    try:
        # Prevent SKU changes on deleted items (items with "-deleted" or "-deleted-{id}" pattern)
        if sku and sku != item['sku']:
            # Check if current item is deleted (has "-deleted-" pattern or ends with "-deleted")
            if '-deleted-' in item['sku'] or item['sku'].endswith('-deleted'):
                return jsonify({"error": "لا يمكن تعديل SKU للقطعة المحذوفة"}), 400
            
            # Check if new SKU already exists
            existing = stock_model.get_stock_item_by_sku(sku)
            if existing:
                return jsonify({"error": get_message("duplicate_sku")}), 409

        # Prepare update data
        update_data = {}
        if sku:
            update_data['sku'] = sku
        if name:
            update_data['name'] = name
        if active is not None:
            update_data['active'] = active
        if user_id:
            update_data['updated_by'] = user_id
        if price_customer is not None:
            update_data['price_customer'] = price_customer
        if price_merchant is not None:
            update_data['price_merchant'] = price_merchant

        # Perform update
        success = stock_model.update_stock_item(item_id, update_data)
        
        if success:
            # Return updated item
            updated_item = stock_model.get_stock_item_by_id(item_id)
            if updated_item['type'] == 'product':
                updated_item['components'] = stock_model.get_product_components(item_id)
            return jsonify(updated_item)
        else:
            return jsonify({"error": get_message("no_changes")}), 400

    except Exception as e:
        return jsonify({"error": get_message("failed_update")}), 500

@stock_api_blueprint.route('/items/<int:item_id>', methods=['DELETE'])
@require_auth
def delete_item(item_id):
    """Delete a stock item (part or product)."""
    try:
        # Check if item exists
        item = stock_model.get_stock_item_by_id(item_id)
        if not item:
            return jsonify({"error": get_message("not_found_stock")}), 404
        
        # Check for dependencies (only active service tickets block deletion)
        dependencies = stock_model.check_item_dependencies(item_id)
        if dependencies['has_dependencies']:
            # Build Arabic error message with ticket details (only active tickets block)
            if dependencies['active_service_items_count'] > 0:
                ticket_list = []
                for ticket in dependencies['service_tickets']:
                    ticket_list.append(f"{ticket['ticket_number']} ({ticket['status']})")
                ticket_str = '، '.join(ticket_list)
                error_msg = f"لا يمكن حذف القطعة: القطعة مرتبطة بـ {dependencies['active_service_items_count']} عنصر خدمة في تذاكر: {ticket_str}"
            else:
                error_msg = "لا يمكن حذف القطعة"
            
            response_data = {
                "error": error_msg,
                "details": {
                    "service_items_count": dependencies['active_service_items_count'],
                    "stock_movements_count": dependencies['stock_movements_count'],
                    "service_tickets": dependencies['service_tickets']
                }
            }
            return jsonify(response_data), 409
        
        # Perform soft deletion (sets active = FALSE and appends "-deleted-{id}" to SKU)
        try:
            success = stock_model.delete_stock_item(item_id)
            if success:
                return jsonify({
                    "message": get_message("deleted"),
                    "item_id": item_id
                }), 200
            else:
                return jsonify({"error": get_message("failed_delete")}), 500
        except Exception as delete_error:
            # Log the actual error for debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error deleting stock item {item_id}: {str(delete_error)}", exc_info=True)
            return jsonify({
                "error": get_message("failed_delete"),
                "details": str(delete_error) if logger.level <= logging.DEBUG else None
            }), 500
            
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Unexpected error in delete_item endpoint: {str(e)}", exc_info=True)
        return jsonify({"error": get_message("failed_delete")}), 500

@stock_api_blueprint.route('/items/<int:item_id>/adjust', methods=['POST'])
@require_auth
def adjust_item_quantity(item_id):
    """Adjust stock for an item."""
    data = request.get_json()
    quantity_delta = data.get('quantity_delta')
    reason = data.get('reason')
    user_id = data.get('user_id')

    if not all([quantity_delta, reason, user_id]):
        return jsonify({"error": get_message("missing_fields")}), 400

    try:
        success = stock_manager.adjust_stock(item_id, quantity_delta, reason, user_id)
        if success:
            updated_item = stock_model.get_stock_item_by_id(item_id)
            return jsonify(updated_item)
    except stock_manager.StockManagerException as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"error": get_message("failed_update")}), 500

@stock_api_blueprint.route('/manual', methods=['POST'])
@require_auth
def manual():
    """Manually adjust stock."""
    data = request.get_json()
    sku = data.get('sku')
    quantity = data.get('quantity')
    condition = data.get('condition')
    user_id = data.get('user_id')
    ticket_id = data.get('ticket_id')
    notes = data.get('notes', '')

    # Validate required fields
    if not all([sku, quantity is not None, condition, user_id]):
        return jsonify({
            "error": get_message("missing_fields")
        }), 400

    # Validate condition
    if condition not in ['valid', 'damaged']:
        return jsonify({"error": get_message("invalid_condition")}), 400

    try:
        # Look up item by SKU
        item = stock_model.get_stock_item_by_sku(sku)
        if not item:
            return jsonify({"error": get_message("not_found_stock")}), 404

        # Perform manual adjustment
        success = stock_manager.manual_stock_adjustment(
            item['id'], quantity, condition, user_id, ticket_id, notes
        )
        
        if success:
            # Return updated item details
            updated_item = stock_model.get_stock_item_by_id(item['id'])
            return jsonify({
                "message": get_message("stock_adjusted"),
                "item": updated_item
            })

    except stock_manager.StockManagerException as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": get_message("failed_update")}), 500

@stock_api_blueprint.route('/items/<int:product_id>/components', methods=['POST'])
@require_auth
def add_component(product_id):
    """Add a component to a product."""
    data = request.get_json()
    part_id = data.get('part_id')
    quantity_needed = data.get('quantity_needed')

    if not all([part_id, quantity_needed]):
        return jsonify({"error": get_message("missing_fields")}), 400

    try:
        # Verify product exists and is a product
        product = stock_model.get_stock_item_by_id(product_id)
        if not product:
            return jsonify({"error": get_message("not_found_stock")}), 404
        
        if product['type'] != 'product':
            return jsonify({"error": get_message("invalid_item_type")}), 400

        # Add component using helper function
        _add_component_to_product(product_id, part_id, quantity_needed)

        # Return updated components
        components = stock_model.get_product_components(product_id)
        return jsonify({"components": components}), 201

    except ValueError as e:
        return jsonify({"error": get_message("not_found_item")}), 404
    except Exception as e:
        return jsonify({"error": get_message("failed_update")}), 500

@stock_api_blueprint.route('/items/<int:product_id>/components/<int:component_id>', methods=['DELETE'])
@require_auth
def remove_component(product_id, component_id):
    """Remove a component from a product."""
    try:
        # Verify product exists
        product = stock_model.get_stock_item_by_id(product_id)
        if not product:
            return jsonify({"error": get_message("not_found_stock")}), 404
        
        # Delete the component
        delete_sql = "DELETE FROM product_components WHERE id = %s AND product_id = %s"
        execute_update(delete_sql, (component_id, product_id))
        
        # Return updated components
        components = stock_model.get_product_components(product_id)
        return jsonify({"components": components})

    except Exception as e:
        return jsonify({"error": get_message("failed_delete")}), 500

@stock_api_blueprint.route('/items/<int:item_id>', methods=['GET'])
@require_auth
def get_item(item_id):
    """Get stock item details."""
    item = stock_model.get_stock_item_by_id(item_id)
    if not item:
        return jsonify({"error": get_message("not_found_stock")}), 404
    
    # Include components if it's a product
    if item['type'] == 'product':
        item['components'] = stock_model.get_product_components(item_id)
    
    return jsonify(item)

@stock_api_blueprint.route('/items', methods=['GET'])
@require_auth
def list_items():
    """List all stock items, with optional filtering by type.
    Only returns active items by default (deleted items are hidden). 
    Use active_only=false to include deleted items (for admin/debug purposes).
    If limit parameter is provided, it is ignored and all items are returned.
    Note: Deleted items still appear in stock movements/history (they reference by ID).
    """
    try:
        item_type = request.args.get('type') # e.g., 'part', 'product'
        limit_param = request.args.get('limit', type=int)
        # If limit is provided, ignore it and return all items
        limit = None
        offset = request.args.get('offset', 0, type=int)
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        
        items = stock_model.get_stock_items_by_type(item_type, limit, offset, active_only=active_only)
        return jsonify(items)
    
    except ConnectionError as e:
        logger.error(f"Database connection error in list_items: {str(e)}")
        return jsonify({
            "error": "Database connection failed",
            "message": "Unable to connect to the database. Please ensure MySQL is running.",
            "details": str(e)
        }), 503  # Service Unavailable
    
    except Exception as e:
        logger.error(f"Unexpected error in list_items: {str(e)}", exc_info=True)
        return jsonify({
            "error": "Internal server error",
            "message": "An unexpected error occurred while fetching stock items.",
            "details": str(e)
        }), 500

@stock_api_blueprint.route('/movements', methods=['GET'])
@require_auth
def list_movements():
    """Get stock movement history with filtering and pagination."""
    # Parse query parameters
    item_id = request.args.get('item_id', type=int)
    movement_type = request.args.get('movement_type')
    reference_type = request.args.get('reference_type')
    reference_id = request.args.get('reference_id', type=int)
    created_by = request.args.get('created_by')
    condition = request.args.get('condition')
    item_type = request.args.get('item_type')
    service_type = request.args.get('service_type')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    limit = request.args.get('limit', 50, type=int)
    offset = request.args.get('offset', 0, type=int)
    order_by = request.args.get('order_by', 'created_at')
    order_direction = request.args.get('order_direction', 'DESC')
    
    # Validate parameters
    if limit < 1 or limit > 100:
        limit = 50
    if offset < 0:
        offset = 0
    if order_direction.upper() not in ['ASC', 'DESC']:
        order_direction = 'DESC'
    if order_by not in ['id', 'created_at', 'movement_type', 'quantity']:
        order_by = 'created_at'
    
    # Validate condition if provided
    if condition and condition not in ['valid', 'damaged']:
        return jsonify({"error": get_message("invalid_condition")}), 400
    
    # Validate item_type if provided
    if item_type and item_type not in ['product', 'part']:
        return jsonify({"error": get_message("invalid_item_type")}), 400
    
    # Validate and handle multiple service types (comma-separated)
    if service_type:
        if ',' in service_type:
            service_types = [st.strip() for st in service_type.split(',') if st.strip()]
            for st in service_types:
                if st not in ['replacement', 'maintenance', 'return']:
                    return jsonify({"error": get_message("invalid_ticket_type")}), 400
            service_type = service_types
        else:
            if service_type not in ['replacement', 'maintenance', 'return']:
                return jsonify({"error": get_message("invalid_ticket_type")}), 400
    
    # Validate and format dates (YYYY-MM-DD format)
    if start_date:
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({"error": get_message("err_invalid_date_format")}), 400
    
    if end_date:
        try:
            datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({"error": get_message("err_invalid_date_format")}), 400
    
    # Handle multiple movement types (comma-separated)
    if movement_type and ',' in movement_type:
        movement_type = [mt.strip().upper() for mt in movement_type.split(',') if mt.strip()]
    elif movement_type:
        movement_type = movement_type.upper()
    
    try:
        # Get movements
        movements = stock_model.get_stock_movements(
            item_id=item_id,
            movement_type=movement_type,
            reference_type=reference_type,
            reference_id=reference_id,
            created_by=created_by,
            condition=condition,
            item_type=item_type,
            service_type=service_type,
            start_date=start_date,
            end_date=end_date,
            limit=limit,
            offset=offset,
            order_by=order_by,
            order_direction=order_direction
        )
        
        # Get total count for pagination
        total_count = stock_model.get_stock_movements_count(
            item_id=item_id,
            movement_type=movement_type,
            reference_type=reference_type,
            reference_id=reference_id,
            created_by=created_by,
            condition=condition,
            item_type=item_type,
            service_type=service_type,
            start_date=start_date,
            end_date=end_date
        )
        
        # Format response with pagination
        response = {
            "data": movements,
            "pagination": {
                "total": total_count,
                "limit": limit,
                "offset": offset,
                "has_more": offset + limit < total_count
            }
        }
        
        return jsonify(response)

    except ConnectionError as e:
        logger.error(f"Database connection error in list_movements: {str(e)}")
        return jsonify({
            "error": "Database connection failed",
            "message": "Unable to connect to the database. Please ensure MySQL is running.",
            "details": str(e),
            "data": [],
            "pagination": {
                "total": 0,
                "limit": limit,
                "offset": offset,
                "has_more": False
            }
        }), 503  # Service Unavailable
    
    except Exception as e:
        logger.error(f"Unexpected error in list_movements: {str(e)}", exc_info=True)
        return jsonify({
            "error": "Internal server error",
            "message": "An unexpected error occurred while fetching stock movements.",
            "details": str(e),
            "data": [],
            "pagination": {
                "total": 0,
                "limit": limit,
                "offset": offset,
                "has_more": False
            }
        }), 500

@stock_api_blueprint.route('/export', methods=['GET'])
@require_auth
def export_stock_items():
    """
    Export stock items to Excel file.
    Query params:
    - item_type: 'product' or 'part' (optional, exports all if not specified)
    """
    try:
        item_type = request.args.get('item_type')

        # Fetch all items based on type (only active items, exclude deleted)
        if item_type:
            items = stock_model.get_stock_items_by_type(item_type=item_type, active_only=True)
        else:
            items = stock_model.get_stock_items_by_type(active_only=True)

        if not items:
            return jsonify({"error": "لا توجد بيانات للتصدير"}), 404

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "المخزون"

        # Define header style
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12, name='Arial')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Arabic headers (right to left)
        headers = ['الحالة', 'سليم', 'تالف', 'محجوز', 'في المخزن', 'اسم الصنف', 'SKU', 'النوع']

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Write data
        for row_num, item in enumerate(items, 2):
            # Calculate valid stock (on_hand - reserved)
            quantity_on_hand = item.get('quantity_on_hand', 0)
            quantity_reserved = item.get('quantity_reserved', 0)
            quantity_damaged = item.get('quantity_damaged', 0)
            valid_stock = quantity_on_hand - quantity_reserved

            # Determine status
            status = 'نشط' if item.get('active', True) else 'غير نشط'
            if not item.get('active', True):
                status = 'محذوف'
            elif quantity_on_hand <= 0:
                status = 'نفد المخزون'
            elif valid_stock <= 0:
                status = 'محجوز بالكامل'

            # Item type in Arabic
            item_type_ar = 'منتج' if item.get('type') == 'product' else 'قطعة'

            # Write row data
            row_data = [
                status,                    # الحالة
                valid_stock,               # سليم (valid stock)
                quantity_damaged,          # تالف
                quantity_reserved,         # محجوز
                quantity_on_hand,          # في المخزن
                item.get('name', ''),      # اسم الصنف
                item.get('sku', ''),       # SKU
                item_type_ar               # النوع
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Color coding for status
                if col_num == 1:  # Status column
                    if status == 'نفد المخزون' or status == 'محذوف':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        cell.font = Font(color="CC0000", bold=True)
                    elif status == 'محجوز بالكامل':
                        cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
                        cell.font = Font(color="FF8C00", bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 15  # الحالة
        ws.column_dimensions['B'].width = 12  # سليم
        ws.column_dimensions['C'].width = 12  # تالف
        ws.column_dimensions['D'].width = 12  # محجوز
        ws.column_dimensions['E'].width = 12  # في المخزن
        ws.column_dimensions['F'].width = 35  # اسم الصنف
        ws.column_dimensions['G'].width = 20  # SKU
        ws.column_dimensions['H'].width = 12  # النوع

        # Set right-to-left direction for the sheet
        ws.sheet_view.rightToLeft = False

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"stock_export_{timestamp}.xlsx"
        if item_type:
            type_ar = 'منتجات' if item_type == 'product' else 'قطع'
            filename = f"stock_{type_ar}_{timestamp}.xlsx"

        # Return file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Export error: {str(e)}")
        return jsonify({"error": "فشل في تصدير البيانات"}), 500
